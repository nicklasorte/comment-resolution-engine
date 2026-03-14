from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from .errors import CREError, ErrorCategory

from .config import ColumnMappingConfig, normalize_header
from .spreadsheet_contract import CANONICAL_INTERNAL_ORDER, CANONICAL_SPREADSHEET_HEADERS, require_canonical_headers, reorder_to_canonical

if TYPE_CHECKING:
    import pandas as pd

OPTIONAL_INTERNAL_COLUMNS = [
    "revision",
    "resolved_against_revision",
    "line_number",
    "wg_chain_comments",
    "status",
    "disposition",
    "report_context",
    "resolution_task",
    "generation_mode",
    "rule_id",
    "rule_source",
    "rule_version",
    "rules_profile",
    "rules_version",
    "matched_rule_types",
    "review_status",
    "confidence_score",
    "provenance_record_id",
    "comment_cluster_id",
    "intent_classification",
    "section_group",
    "heat_level",
    "validation_status",
    "validation_notes",
    "validation_code",
    "patch_text",
    "patch_source",
    "patch_confidence",
    "resolution_basis",
    "context_confidence",
    "cluster_label",
    "cluster_size",
    "shared_resolution_id",
    "canonical_term_used",
]


def _require_pandas():
    try:
        import pandas as pd
    except ModuleNotFoundError as exc:
        raise CREError(ErrorCategory.EXTRACTION_ERROR, "pandas is required for Excel processing. Install dependencies with `pip install -r requirements.txt`.") from exc
    return pd


def _require_openpyxl():
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise CREError(ErrorCategory.EXTRACTION_ERROR, "openpyxl is required for Excel formatting. Install dependencies with `pip install -r requirements.txt`.") from exc
    return load_workbook, Alignment, Font, get_column_letter


def _build_header_lookup(columns: list[str]) -> dict[str, str]:
    return {normalize_header(col): col for col in columns}


def normalize_comment_matrix(df: "pd.DataFrame", mapping: ColumnMappingConfig) -> "pd.DataFrame":
    pd = _require_pandas()
    require_canonical_headers(df.columns.tolist())
    lookup = _build_header_lookup(df.columns.tolist())
    normalized = {}

    for canonical in [*CANONICAL_INTERNAL_ORDER, *OPTIONAL_INTERNAL_COLUMNS]:
        raw_column_name = ""
        for variant in mapping.all_variants(canonical):
            if variant in lookup:
                raw_column_name = lookup[variant]
                break
        normalized[canonical] = df[raw_column_name] if raw_column_name else ""

    out_df = pd.DataFrame(normalized)
    return out_df.fillna("")


def read_comment_matrix(path: str | Path, mapping: ColumnMappingConfig):
    pd = _require_pandas()
    df = pd.read_excel(path)
    return normalize_comment_matrix(df, mapping)


def write_resolution_workbook(df, output_path: str | Path, include_metadata: bool = False) -> None:
    load_workbook, Alignment, Font, get_column_letter = _require_openpyxl()

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    ordered = reorder_to_canonical(df.copy(), include_metadata=include_metadata)
    ordered.to_excel(output, index=False)

    wb = load_workbook(output)
    ws = wb.active
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    width_by_header = {
        "Comment Number": 16,
        "Reviewer Initials": 16,
        "Agency": 18,
        "Report Version": 18,
        "Section": 16,
        "Page": 12,
        "Line": 14,
        "Comment Type: Editorial/Grammar, Clarification, Technical": 26,
        "Agency Notes": 55,
        "Agency Suggested Text Change": 55,
        "NTIA Comments": 65,
        "Comment Disposition": 18,
        "Resolution": 70,
    }

    wrap_headers = {
        "Agency Notes",
        "Agency Suggested Text Change",
        "NTIA Comments",
        "Resolution",
    }

    for idx, col_name in enumerate(ordered.columns, start=1):
        width = width_by_header.get(col_name, 18)
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row == 1:
                cell.font = Font(bold=True)
            header_val = ws.cell(row=1, column=cell.column).value
            cell.alignment = Alignment(vertical="top", wrap_text=str(header_val) in wrap_headers)

    wb.save(output)
