#!/usr/bin/env python3
"""
resolve_comments.py

MVP execution pipeline: adjudicate a comment resolution matrix against a
revised working paper PDF and output an updated matrix.

Inputs:
  1. Canonical comment resolution matrix spreadsheet (.xlsx or .csv)
  2. Revised working paper PDF

Output:
  adjudicated_comment_matrix.xlsx  (or path specified by --output)

Usage:
    python resolve_comments.py \\
        --matrix input_matrix.xlsx \\
        --paper working_paper_rev2.pdf \\
        --output adjudicated_matrix.xlsx
"""

from __future__ import annotations

import argparse
import importlib
import sys
from pathlib import Path
from typing import NoReturn

# ---------------------------------------------------------------------------
# Dependency helpers
# ---------------------------------------------------------------------------

_INSTALL_HINT = (
    "Create a virtualenv and install dependencies:\n"
    "  python3 -m venv .venv && source .venv/bin/activate && pip install -r requirements.txt"
)


def _exit_dependency_error(message: str) -> NoReturn:
    print(f"{message}\n{_INSTALL_HINT}")
    sys.exit(1)


def _require_module(module_name: str, friendly_name: str):
    try:
        return importlib.import_module(module_name)
    except ModuleNotFoundError as exc:
        _exit_dependency_error(f"ERROR: {friendly_name} is required for this command. ({exc})")
    except Exception as exc:  # Handles compiled extension issues like missing _cffi_backend
        _exit_dependency_error(
            f"ERROR: {friendly_name} failed to import ({type(exc).__name__}: {exc}). "
            "Reinstall dependencies."
        )


def _require_pandas():
    return _require_module("pandas", "pandas")


def _require_openpyxl():
    _require_module("openpyxl", "openpyxl")


def _require_pypdf():
    return _require_module("pypdf", "pypdf/cryptography")


def preflight_check(verbose: bool = False) -> None:
    """
    Validate that all required runtime dependencies can be imported.
    """
    for mod, friendly in (
        ("pandas", "pandas"),
        ("openpyxl", "openpyxl"),
        ("pypdf", "pypdf"),
        ("cryptography", "cryptography"),
    ):
        _require_module(mod, friendly)
    if verbose:
        print("Preflight OK: required imports succeeded.")


# ---------------------------------------------------------------------------
# Spreadsheet contract (re-uses canonical definitions from the engine)
# ---------------------------------------------------------------------------

def _engine_src() -> Path:
    return Path(__file__).parent / "src"


def _canonical_headers() -> list[str]:
    sys.path.insert(0, str(_engine_src()))
    from comment_resolution_engine.spreadsheet_contract import CANONICAL_SPREADSHEET_HEADERS
    return list(CANONICAL_SPREADSHEET_HEADERS)


def _validate_headers(columns: list[str]) -> None:
    sys.path.insert(0, str(_engine_src()))
    from comment_resolution_engine.spreadsheet_contract import require_canonical_headers
    require_canonical_headers(columns)


# ---------------------------------------------------------------------------
# Load matrix
# ---------------------------------------------------------------------------

def load_matrix(path: str):
    pd = _require_pandas()
    p = Path(path)
    if p.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    return df


# ---------------------------------------------------------------------------
# PDF parsing
# ---------------------------------------------------------------------------

def parse_pdf(path: str) -> str:
    """Extract full text from the working paper PDF."""
    try:
        pypdf = _require_pypdf()
        PdfReader = pypdf.PdfReader
    except SystemExit:
        raise
    except Exception as exc:
        _exit_dependency_error(
            f"ERROR: PDF parsing dependencies failed to load ({type(exc).__name__}: {exc}). "
            "Reinstall dependencies."
        )

    try:
        reader = PdfReader(path)
        pages = [page.extract_text() or "" for page in reader.pages]
        return "\n".join(pages)
    except Exception as exc:
        print(
            f"ERROR: Could not parse PDF '{path}': {type(exc).__name__}: {exc}\n"
            "Hint: ensure the PDF is not password protected and dependencies are installed "
            "with `pip install -r requirements.txt`."
        )
        sys.exit(1)


# ---------------------------------------------------------------------------
# Per-row adjudication logic
# ---------------------------------------------------------------------------

_COMPLETED_VALUES = {"completed", "complete", "closed", "resolved", "done"}


def _is_completed(row, status_col: str | None) -> bool:
    """Return True if the row is already marked as completed/resolved."""
    if status_col and status_col in row.index:
        val = str(row[status_col]).strip().lower()
        if val in _COMPLETED_VALUES:
            return True
    disp = str(row.get("Comment Disposition", "")).strip().lower()
    return disp in _COMPLETED_VALUES


def _clean(value) -> str:
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", ""} else text


def _find_revision_reference(row) -> str:
    """Build a section/line revision reference string from row metadata."""
    parts: list[str] = []
    section = _clean(row.get("Section", ""))
    page = _clean(row.get("Page", ""))
    line = _clean(row.get("Line", ""))
    if section:
        parts.append(f"Section {section}")
    if page:
        parts.append(f"p. {page}")
    if line:
        parts.append(f"l. {line}")
    return ", ".join(parts)


def _determine_disposition(row) -> str:
    """
    Heuristic disposition based on comment type and content.

    Returns one of: Accepted | Partially Accepted | Rejected
    """
    comment_type = _clean(
        row.get("Comment Type: Editorial/Grammar, Clarification, Technical", "")
    ).lower()
    agency_notes = _clean(row.get("Agency Notes", "")).lower()
    suggested = _clean(row.get("Agency Suggested Text Change", ""))

    # Explicit reject signals in the agency notes
    reject_signals = {
        "out of scope",
        "not applicable",
        "already addressed",
        "already reflected",
        "already included",
    }
    if any(sig in agency_notes for sig in reject_signals):
        return "Rejected"

    # Editorial/grammar items are accepted by default
    if "editorial" in comment_type or "grammar" in comment_type:
        return "Accepted"

    # Technical items without suggested text → Partially Accepted
    if "technical" in comment_type and not suggested:
        return "Partially Accepted"

    return "Accepted"


def _generate_response(row, disposition: str) -> str:
    """Generate a plain-language response to the agency comment."""
    agency_notes = _clean(row.get("Agency Notes", ""))
    suggested = _clean(row.get("Agency Suggested Text Change", ""))
    truncated_notes = agency_notes[:150] + ("..." if len(agency_notes) > 150 else "")

    if disposition == "Rejected":
        return (
            f"The comment has been reviewed. The existing working paper text is retained. "
            f"The concern — '{truncated_notes}' — does not require a revision at this stage."
        )
    if disposition == "Partially Accepted":
        base = (
            f"The comment is acknowledged. Clarifications will be incorporated to address: "
            f"'{truncated_notes}'."
        )
        if suggested:
            base += " The suggested text was considered but only partially adopted."
        return base
    # Accepted
    if suggested:
        truncated_suggestion = suggested[:200] + ("..." if len(suggested) > 200 else "")
        return (
            f"The comment is accepted. The working paper will be updated to reflect: "
            f"'{truncated_suggestion}'."
        )
    return (
        f"The comment is accepted. The working paper will be updated to address: "
        f"'{truncated_notes}'."
    )


def _generate_notes(row, disposition: str) -> str:
    """Generate optional explanatory notes for the adjudication."""
    comment_type = _clean(
        row.get("Comment Type: Editorial/Grammar, Clarification, Technical", "")
    )
    section = _clean(row.get("Section", ""))
    parts: list[str] = []
    if comment_type:
        parts.append(f"Type: {comment_type}")
    if disposition == "Rejected" and section:
        parts.append(f"No change to Section {section}.")
    elif disposition == "Partially Accepted":
        parts.append("Partial revision pending editorial review.")
    return " | ".join(parts)


def adjudicate_row(
    row,
    status_col: str | None,
) -> tuple[str, str, str, str]:
    """
    Adjudicate a single comment row.

    Returns:
        (disposition, response, revision_reference, notes)
    """
    if _is_completed(row, status_col):
        revision_ref = _find_revision_reference(row)
        # Preserve any existing resolution text; do not generate a new response.
        existing_resolution = _clean(row.get("Resolution", ""))
        return "Completed", existing_resolution, revision_ref, ""

    disposition = _determine_disposition(row)
    response = _generate_response(row, disposition)
    revision_ref = _find_revision_reference(row)
    notes = _generate_notes(row, disposition)
    return disposition, response, revision_ref, notes


# ---------------------------------------------------------------------------
# Output formatting
# ---------------------------------------------------------------------------

def _format_ntia_comments(revision_ref: str, notes: str) -> str:
    """Combine revision reference and notes into the NTIA Comments field."""
    parts: list[str] = []
    if revision_ref:
        parts.append(f"Ref: {revision_ref}")
    if notes:
        parts.append(notes)
    return " | ".join(parts)


def write_output(df, output_path: str) -> None:
    """Write the adjudicated DataFrame to a formatted Excel workbook."""
    _require_openpyxl()
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(out, index=False)

    wb = load_workbook(out)
    ws = wb.active
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    width_hints: dict[str, int] = {
        "Comment Number": 16,
        "Reviewer Initials": 16,
        "Agency": 18,
        "Report Version": 18,
        "Section": 16,
        "Page": 12,
        "Line": 14,
        "Comment Type: Editorial/Grammar, Clarification, Technical": 26,
        "Agency Notes": 55,
        "Agency Suggested Text Change": 55,
        "NTIA Comments": 55,
        "Comment Disposition": 22,
        "Resolution": 70,
    }
    wrap_cols = {
        "Agency Notes",
        "Agency Suggested Text Change",
        "NTIA Comments",
        "Resolution",
    }

    for idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width_hints.get(col_name, 20)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            header_val = ws.cell(row=1, column=cell.column).value or ""
            if cell.row == 1:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=str(header_val) in wrap_cols,
            )

    wb.save(out)


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def run(matrix_path: str, paper_path: str, output_path: str) -> None:
    preflight_check()

    print(f"Loading matrix: {matrix_path}")
    df = load_matrix(matrix_path)

    print("Validating canonical headers...")
    _validate_headers(df.columns.tolist())

    print(f"Parsing working paper PDF: {paper_path}")
    paper_text = parse_pdf(paper_path)
    if paper_text:
        print(f"  Extracted {len(paper_text):,} characters from PDF.")
    else:
        print("  No PDF text extracted; proceeding without PDF context.")

    # Detect optional STATUS column (case-insensitive)
    status_col: str | None = next(
        (col for col in df.columns if col.strip().lower() == "status"),
        None,
    )
    if status_col:
        print(f"  Found STATUS column: '{status_col}'")

    dispositions: list[str] = []
    responses: list[str] = []
    ntia_values: list[str] = []

    print(f"Adjudicating {len(df)} comment row(s)...")
    for _, row in df.iterrows():
        disposition, response, revision_ref, notes = adjudicate_row(row, status_col)
        dispositions.append(disposition)
        responses.append(response)
        ntia_values.append(_format_ntia_comments(revision_ref, notes))

    # Write results into the canonical columns
    df["Comment Disposition"] = dispositions
    df["Resolution"] = responses
    df["NTIA Comments"] = ntia_values

    # Re-order: canonical headers first, then any extra columns
    canonical = _canonical_headers()
    ordered_canonical = [h for h in canonical if h in df.columns]
    extra = [c for c in df.columns if c not in set(ordered_canonical)]
    df = df[ordered_canonical + extra]

    print(f"Writing output: {output_path}")
    write_output(df, output_path)
    print(f"Done. Wrote {len(df)} row(s) to {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Adjudicate a comment resolution matrix against a revised working paper PDF "
            "and output an updated matrix."
        )
    )
    parser.add_argument(
        "--matrix",
        required=True,
        help="Path to the input comment resolution matrix (.xlsx or .csv).",
    )
    parser.add_argument(
        "--paper",
        required=True,
        help="Path to the revised working paper PDF.",
    )
    parser.add_argument(
        "--output",
        default="adjudicated_comment_matrix.xlsx",
        help="Output path for the adjudicated matrix (default: adjudicated_comment_matrix.xlsx).",
    )
    parser.add_argument(
        "--preflight",
        action="store_true",
        help="Validate required dependencies and exit.",
    )
    args = parser.parse_args()

    if args.preflight:
        preflight_check(verbose=True)
        return

    if not Path(args.matrix).exists():
        print(f"ERROR: Matrix file not found: {args.matrix}")
        sys.exit(1)
    if not Path(args.paper).exists():
        print(f"ERROR: Working paper PDF not found: {args.paper}")
        sys.exit(1)

    run(args.matrix, args.paper, args.output)


if __name__ == "__main__":
    main()
