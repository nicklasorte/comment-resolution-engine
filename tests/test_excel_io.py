from pathlib import Path

import pytest

pd = pytest.importorskip("pandas")
openpyxl = pytest.importorskip("openpyxl")
from openpyxl import load_workbook

from comment_resolution_engine.excel_io import write_resolution_workbook


def test_write_resolution_workbook_formats_sheet(tmp_path: Path):
    output = tmp_path / "out.xlsx"
    df = pd.DataFrame(
        {
            "Comment Number": ["1"],
            "Reviewer Initials": ["AB"],
            "Agency": ["Agency X"],
            "Report Version": ["Draft"],
            "Section": ["2.1"],
            "Page": ["10"],
            "Line": ["33"],
            "Comment Type": ["Clarification"],
            "Agency Notes": ["Clarify assumptions"],
            "Agency Suggested Text Change": ["The report clarifies the scope of assumptions."],
            "NTIA Comments": ["Accept. Issue noted and agency suggested text is incorporated with NTIA edits as needed."],
            "Comment Disposition": ["Accept"],
            "Resolution": ["The report clarifies the scope of assumptions."],
            "Report Context": ["L33: System assumptions are defined here"],
            "Resolution Task": ["Draft NTIA Comments, Comment Disposition (Accept/Reject), and Resolution text that can be inserted into the report. Comment number: 1."],
        }
    )

    write_resolution_workbook(df, output)

    wb = load_workbook(output)
    ws = wb.active
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    assert ws["A1"].value == "Comment Number"

    headers = [cell.value for cell in ws[1]]
    ntia_idx = headers.index("NTIA Comments") + 1
    resolution_idx = headers.index("Resolution") + 1
    assert ws.cell(row=2, column=ntia_idx).alignment.wrap_text is True
    assert ws.cell(row=2, column=resolution_idx).alignment.wrap_text is True
